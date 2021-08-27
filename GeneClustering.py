import openpyxl

minimum = input("Enter minimum expression percentage (between 0 and 100): ")
maximum = input("Enter maximum expression percentage (between 0 and 100): ")
name = input("Enter name for file that will store results: ")

#Change to file path of excel file
filePath = ("C:\\Users\\areeb\\Desktop\\HeLa.SE.expression.merge.xlsx")

#Object used to extract data from excel file
workbook = openpyxl.load_workbook(filePath)
sheet = workbook.active

#Dimensions of gene expression matrix
columnCount = sheet.max_column
rowCount = sheet.max_row

geneExpression = {}

for columnNumber in range(2, columnCount):
    for rowNumber in range(2, rowCount):
        geneName = sheet.cell(row=rowNumber, column = 1)
        singleCellName = sheet.cell(row=1, column=columnNumber)
        datum = sheet.cell(row=rowNumber, column=columnNumber)
        mapPair = (geneName.value, singleCellName.value)

geneExpressionCount = {}
geneExpressionTotal = {}
geneExpressionPercentage = {}

for key in geneExpression:
    if (key[0] not in geneExpressionCount):
        geneExpressionCount[key[0]] = 0
        geneExpressionTotal[key[0]] = 1
    if (geneExpression[key] > 0):
        geneExpressionCount[key[0]] = geneExpressionCount[key[0]] + 1
        geneExpressionTotal[key[0]] = geneExpressionTotal[key[0]] + 1
    else:
        geneExpressionTotal[key[0]] = geneExpressionTotal[key[0]] + 1



for key in geneExpressionCount:
    geneExpressionPercentage[key] = (float(geneExpressionCount[key]) / geneExpressionTotal[key]) * 100.0


repeat = True

while repeat:
    storageFileName = name + "_" + minimum + "_to_" + maximum + ".txt"
    storageFile = open(storageFileName, 'w')
    for key in geneExpressionPercentage:
        geneName = str(key)
        percentageExpression = geneExpressionPercentage[key]
        if (percentageExpression >= float(minimum) and percentageExpression <= float(maximum)):
            percentageExpressionString = str(percentageExpression)
            storageFile.write("Gene " + geneName + " is expressed " + percentageExpressionString + " percent of the time" + "\n")

    repeatString = input("Select new range for expression percentage? Enter Y or N: ")
    if (repeatString == "Y"):
        minimum = input("Enter minimum expression percentage (between 0 and 100): ")
        maximum = input("Enter maximum expression percentage (between 0 and 100): ")
    else:
        repeat = False

print("Your data has been stored in text files")
    

    
















